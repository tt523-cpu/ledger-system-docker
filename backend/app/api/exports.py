from datetime import date, timedelta
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from sqlalchemy import extract, func, select
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import get_accessible_platform_ids, get_current_tenant_id, require_module, require_roles
from app.core.time_utils import beijing_now
from app.models.entities import AccountSnapshot, AuditLog, Category, DailySummary, PaymentMethod, Platform, ReportExcelShare, Shift, Tenant, TenantPlatformAccess, Transaction, User
from app.models.enums import UserRole


router = APIRouter(prefix="/exports", tags=["exports"])
PUBLIC_ROUTER = APIRouter(prefix="/public", tags=["public"])
SHARED_REPORT_DIR = Path("shared_reports")
EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def apply_money_format(ws, cols: list[int], start_row: int = 2):
    for r in range(start_row, ws.max_row + 1):
        for c in cols:
            ws.cell(row=r, column=c).number_format = "0.00"


THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)
HEADER_FILL = PatternFill("solid", fgColor="F5F7FA")
SECTION_FILL = PatternFill("solid", fgColor="EEF2F7")
WARNING_FILL = PatternFill("solid", fgColor="FFF4E5")


def _money(value) -> float:
    return float(value or 0)


def _style_range(ws, row_start: int, row_end: int, col_start: int, col_end: int):
    for row in ws.iter_rows(min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _append_section_title(ws, title: str, columns: int = 6):
    ws.append([title])
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
    cell = ws.cell(row=row, column=1)
    cell.fill = SECTION_FILL
    cell.font = Font(bold=True)
    _style_range(ws, row, row, 1, columns)


def _append_table(ws, headers: list[str], rows: list[list], money_cols: list[int] | None = None, summary_rows: set[int] | None = None):
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    end_row = ws.max_row
    _style_range(ws, header_row, end_row, 1, len(headers))
    if money_cols:
        for row_idx in range(header_row + 1, end_row + 1):
            for col in money_cols:
                ws.cell(row=row_idx, column=col).number_format = "0.00"
    if summary_rows:
        for row_idx in summary_rows:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = SECTION_FILL
                ws.cell(row=row_idx, column=col).font = Font(bold=True)
    ws.append([])


def _scoped_platform_ids(db: Session, current_user: User) -> list[int] | None:
    if current_user.role == UserRole.SUPER_ADMIN.value:
        return None
    return get_accessible_platform_ids(db, current_user)


def _tenant_select(model, tenant_id: int | None):
    stmt = select(model)
    if tenant_id is not None:
        stmt = stmt.where(model.tenant_id == tenant_id)
    return stmt


@router.get("/daily-excel")
def export_daily_excel(
    bill_date: str,
    shift_id: int | None = None,
    platform_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("reports.query")),
):
    allowed = _scoped_platform_ids(db, current_user)
    tenant_id = get_current_tenant_id(db, current_user)
    if allowed == []:
        allowed = [-1]
    wb = Workbook()
    ws = wb.active
    ws.title = "日报表"
    ws.append(["日期", "班次", "平台", "充值", "支出", "净营业"])

    stmt = select(DailySummary).where(DailySummary.bill_date == bill_date)
    if allowed is not None:
        stmt = stmt.where(DailySummary.platform_id.in_(allowed))
    if shift_id:
        stmt = stmt.where(DailySummary.shift_id == shift_id)
    if platform_id:
        stmt = stmt.where(DailySummary.platform_id == platform_id)
    rows = db.execute(stmt.order_by(DailySummary.shift_id.asc(), DailySummary.platform_id.asc())).scalars().all()
    platform_stmt = select(Platform.id, Platform.name)
    shift_stmt = select(Shift.id, Shift.name)
    if tenant_id is not None:
        platform_stmt = platform_stmt.where(Platform.tenant_id == tenant_id)
        shift_stmt = shift_stmt.where(Shift.tenant_id == tenant_id)
    platform_map = {r[0]: r[1] for r in db.execute(platform_stmt).all()}
    shift_map = {r[0]: r[1] for r in db.execute(shift_stmt).all()}

    for row in rows:
        ws.append([
            row.bill_date.isoformat(),
            shift_map.get(row.shift_id, row.shift_id),
            platform_map.get(row.platform_id, f"平台#{row.platform_id}"),
            float(row.total_income or 0),
            float(row.total_expense or 0),
            float(row.net_profit or 0),
        ])
    apply_money_format(ws, [4, 5, 6])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"daily-{bill_date}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _build_report_query_excel(
    start_date: str,
    end_date: str,
    shift_id: int | None = None,
    platform_id: int | None = None,
    db: Session | None = None,
    current_user: User | None = None,
) -> tuple[BytesIO, str]:
    assert db is not None
    assert current_user is not None
    allowed = _scoped_platform_ids(db, current_user)
    tenant_id = get_current_tenant_id(db, current_user)
    if allowed == []:
        allowed = [-1]
    wb = Workbook()
    ws = wb.active
    ws.title = "报表查询"
    ws.sheet_view.showGridLines = False
    for col, width in {"A": 18, "B": 18, "C": 26, "D": 18, "E": 34, "F": 18}.items():
        ws.column_dimensions[col].width = width

    stmt = select(DailySummary).where(DailySummary.bill_date >= start_date, DailySummary.bill_date <= end_date)
    if allowed is not None:
        stmt = stmt.where(DailySummary.platform_id.in_(allowed))
    if shift_id:
        stmt = stmt.where(DailySummary.shift_id == shift_id)
    if platform_id:
        stmt = stmt.where(DailySummary.platform_id == platform_id)
    rows = db.execute(stmt.order_by(DailySummary.bill_date.asc(), DailySummary.shift_id.asc(), DailySummary.platform_id.asc())).scalars().all()
    platform_stmt = select(Platform.id, Platform.name)
    shift_stmt = select(Shift.id, Shift.name)
    if tenant_id is not None:
        platform_stmt = platform_stmt.where(Platform.tenant_id == tenant_id)
        shift_stmt = shift_stmt.where(Shift.tenant_id == tenant_id)
    platform_map = {r[0]: r[1] for r in db.execute(platform_stmt).all()}
    shift_map = {r[0]: r[1] for r in db.execute(shift_stmt).all()}

    total_income = sum(_money(r.total_income) for r in rows)
    total_expense = sum(_money(r.total_expense) for r in rows)
    total_net = sum(_money(r.net_profit) for r in rows)

    is_day_report = start_date == end_date
    summary_row = ws.max_row + 1
    ws.append([f"充值合计：{total_income:.2f}", "", f"支出合计：{total_expense:.2f}", "", f"净营业：{total_net:.2f}", ""])
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=2)
    ws.merge_cells(start_row=summary_row, start_column=3, end_row=summary_row, end_column=4)
    ws.merge_cells(start_row=summary_row, start_column=5, end_row=summary_row, end_column=6)
    for col in (1, 3, 5):
        cell = ws.cell(row=summary_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
    _style_range(ws, summary_row, summary_row, 1, 6)
    ws.append([])

    platform_summary_stmt = select(
        DailySummary.platform_id,
        func.sum(DailySummary.total_income),
        func.sum(DailySummary.total_expense),
        func.sum(DailySummary.net_profit),
    ).where(DailySummary.bill_date >= start_date, DailySummary.bill_date <= end_date)
    if allowed is not None:
        platform_summary_stmt = platform_summary_stmt.where(DailySummary.platform_id.in_(allowed))
    if shift_id:
        platform_summary_stmt = platform_summary_stmt.where(DailySummary.shift_id == shift_id)
    if platform_id:
        platform_summary_stmt = platform_summary_stmt.where(DailySummary.platform_id == platform_id)
    platform_summary_rows = db.execute(platform_summary_stmt.group_by(DailySummary.platform_id).order_by(DailySummary.platform_id.asc())).all()

    _append_section_title(ws, "按平台汇总")
    _append_table(
        ws,
        ["平台", "充值", "支出", "净营业"],
        [
            [platform_map.get(r[0], f"平台#{r[0]}"), _money(r[1]), _money(r[2]), _money(r[3])]
            for r in platform_summary_rows
        ],
        money_cols=[2, 3, 4],
    )

    category_map = {c.id: c.name for c in db.execute(_tenant_select(Category, tenant_id)).scalars().all()}
    row_expense_stmt = select(
        Transaction.bill_date,
        Transaction.shift_id,
        Transaction.platform_id,
        Transaction.category_id,
        Transaction.biz_type_label,
        func.sum(Transaction.amount),
    ).where(
        Transaction.bill_date >= start_date,
        Transaction.bill_date <= end_date,
        Transaction.type == "expense",
        Transaction.deleted_at.is_(None),
    )
    if allowed is not None:
        row_expense_stmt = row_expense_stmt.where(Transaction.platform_id.in_(allowed))
    if shift_id:
        row_expense_stmt = row_expense_stmt.where(Transaction.shift_id == shift_id)
    if platform_id:
        row_expense_stmt = row_expense_stmt.where(Transaction.platform_id == platform_id)
    row_expense_rows = db.execute(
        row_expense_stmt.group_by(
            Transaction.bill_date,
            Transaction.shift_id,
            Transaction.platform_id,
            Transaction.category_id,
            Transaction.biz_type_label,
        )
    ).all()
    expense_detail_map = {}
    for bill_date_val, shift_val, platform_val, category_id_val, biz_type_label_val, amount_sum in row_expense_rows:
        key = (bill_date_val.isoformat(), int(shift_val), int(platform_val))
        if category_id_val is None:
            item_name = (biz_type_label_val or "-").strip() or "-"
        else:
            item_name = category_map.get(category_id_val, f"项目#{category_id_val}")
        detail = f"{item_name}:{float(amount_sum or 0):.2f}"
        expense_detail_map.setdefault(key, []).append(detail)

    detail_rows = []
    for row in rows:
        key = (row.bill_date.isoformat(), int(row.shift_id), int(row.platform_id))
        expense_details = "，".join(expense_detail_map.get(key, []))
        expense_cell = f"{_money(row.total_expense):.2f}"
        if expense_details:
            expense_cell = f"{_money(row.total_expense):.2f}（{expense_details}）"
        detail_rows.append(
            [
                row.bill_date.isoformat(),
                shift_map.get(row.shift_id, row.shift_id),
                platform_map.get(row.platform_id, f"平台#{row.platform_id}"),
                _money(row.total_income),
                expense_cell,
                _money(row.net_profit),
            ]
        )
    _append_table(ws, ["日期", "班次", "平台", "充值", "支出", "净营业"], detail_rows, money_cols=[4, 6])

    if not is_day_report:
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"report-{start_date}-to-{end_date}.xlsx"
        return output, filename

    _append_section_title(ws, "账户余额（当日）")
    methods = db.execute(_tenant_select(PaymentMethod, tenant_id).where(PaymentMethod.status == "enabled")).scalars().all()
    total_opening = 0.0
    total_recharge = 0.0
    total_payout = 0.0
    total_closing = 0.0
    balance_rows = []
    channel_subtotals: dict[str, dict[str, float]] = {}
    for pm in methods:
        before_row = db.execute(
            select(func.sum(Transaction.amount), Transaction.type)
            .where(
                Transaction.payment_method_id == pm.id,
                Transaction.bill_date < start_date,
                Transaction.deleted_at.is_(None),
                Transaction.platform_id.in_(allowed) if allowed is not None else True,
            )
            .group_by(Transaction.type)
        ).all()
        opening = float(pm.initial_balance or 0)
        for amount_sum, tx_type in before_row:
            if tx_type == "income":
                opening += float(amount_sum or 0)
            elif tx_type == "expense":
                opening -= float(amount_sum or 0)

        range_row = db.execute(
            select(func.sum(Transaction.amount), Transaction.type)
            .where(
                Transaction.payment_method_id == pm.id,
                Transaction.bill_date >= start_date,
                Transaction.bill_date <= end_date,
                Transaction.deleted_at.is_(None),
                Transaction.platform_id.in_(allowed) if allowed is not None else True,
            )
            .group_by(Transaction.type)
        ).all()
        recharge = 0.0
        payout = 0.0
        for amount_sum, tx_type in range_row:
            if tx_type == "income":
                recharge += _money(amount_sum)
            elif tx_type == "expense":
                payout += _money(amount_sum)
        expense_detail_rows = db.execute(
            select(Transaction.category_id, Transaction.biz_type_label, func.sum(Transaction.amount))
            .where(
                Transaction.payment_method_id == pm.id,
                Transaction.bill_date >= start_date,
                Transaction.bill_date <= end_date,
                Transaction.deleted_at.is_(None),
                Transaction.type == "expense",
                Transaction.platform_id.in_(allowed) if allowed is not None else True,
            )
            .group_by(Transaction.category_id, Transaction.biz_type_label)
        ).all()
        payout_details = []
        for category_id_val, biz_type_label_val, amount_sum in expense_detail_rows:
            if category_id_val is None:
                item_name = (biz_type_label_val or "-").strip() or "-"
            else:
                item_name = category_map.get(category_id_val, f"项目#{category_id_val}")
            payout_details.append(f"{item_name}:{_money(amount_sum):.2f}")
        payout_cell = f"{payout:.2f}"
        if payout_details:
            payout_cell = f"{payout:.2f}（{'，'.join(payout_details)}）"
        closing = opening + recharge - payout
        total_opening += opening
        total_recharge += recharge
        total_payout += payout
        total_closing += closing
        balance_rows.append([pm.name, pm.channel_kind, opening, recharge, payout_cell, closing])

        channel_key = pm.channel_kind or "other"
        if channel_key not in channel_subtotals:
            channel_subtotals[channel_key] = {"opening": 0.0, "recharge": 0.0, "payout": 0.0, "closing": 0.0}
        channel_subtotals[channel_key]["opening"] += opening
        channel_subtotals[channel_key]["recharge"] += recharge
        channel_subtotals[channel_key]["payout"] += payout
        channel_subtotals[channel_key]["closing"] += closing
    balance_rows.append(["总计", "-", total_opening, total_recharge, f"{total_payout:.2f}", total_closing])
    balance_summary_row = ws.max_row + 1 + len(balance_rows)
    _append_table(
        ws,
        ["账户", "通道类型", "期初余额", "充值", "支出", "期末余额"],
        balance_rows,
        money_cols=[3, 4, 6],
        summary_rows={balance_summary_row},
    )

    channel_rows = [
        [channel, values["opening"], values["recharge"], values["payout"], values["closing"]]
        for channel, values in channel_subtotals.items()
    ]
    _append_table(
        ws,
        ["通道类型小计", "期初小计", "充值小计", "支出小计", "期末小计"],
        channel_rows,
        money_cols=[2, 3, 4, 5],
    )

    _append_section_title(ws, "交班营业报表（当日）")
    shift_rows = db.execute(
        select(DailySummary.shift_id, func.sum(DailySummary.total_income), func.sum(DailySummary.total_expense), func.sum(DailySummary.net_profit))
        .where(DailySummary.bill_date == start_date, DailySummary.platform_id.in_(allowed) if allowed is not None else True)
        .group_by(DailySummary.shift_id)
        .order_by(DailySummary.shift_id.asc())
    ).all()
    shift_expense_rows = db.execute(
        select(Transaction.shift_id, Transaction.category_id, Transaction.biz_type_label, func.sum(Transaction.amount))
        .where(
            Transaction.bill_date == start_date,
            Transaction.type == "expense",
            Transaction.deleted_at.is_(None),
            Transaction.platform_id.in_(allowed) if allowed is not None else True,
        )
        .group_by(Transaction.shift_id, Transaction.category_id, Transaction.biz_type_label)
    ).all()
    shift_expense_map: dict[int, list[str]] = {}
    for shift_id_val, category_id_val, biz_type_label_val, amount_sum in shift_expense_rows:
        if category_id_val is None:
            item_name = (biz_type_label_val or "-").strip() or "-"
        else:
            item_name = category_map.get(category_id_val, f"项目#{category_id_val}")
        shift_expense_map.setdefault(int(shift_id_val), []).append(f"{item_name}:{_money(amount_sum):.2f}")
    handover_rows = []
    for r in shift_rows:
        expense = _money(r[2])
        expense_cell = f"{expense:.2f}"
        details = "，".join(shift_expense_map.get(int(r[0]), []))
        if details:
            expense_cell = f"{expense:.2f}（{details}）"
        handover_rows.append([shift_map.get(r[0], r[0]), _money(r[1]), expense_cell, _money(r[3])])
    _append_table(ws, ["班次", "充值", "支出", "营业额"], handover_rows, money_cols=[2, 4])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"report-{start_date}-to-{end_date}.xlsx"
    return output, filename


@router.get("/report-query-excel")
def export_report_query_excel(
    start_date: str,
    end_date: str,
    shift_id: int | None = None,
    platform_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("reports.query")),
):
    output, filename = _build_report_query_excel(
        start_date=start_date,
        end_date=end_date,
        shift_id=shift_id,
        platform_id=platform_id,
        db=db,
        current_user=current_user,
    )
    return StreamingResponse(
        output,
        media_type=EXCEL_MEDIA_TYPE,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/report-query-share-excel")
def share_report_query_excel(
    start_date: str,
    end_date: str,
    shift_id: int | None = None,
    platform_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("reports.query")),
):
    output, filename = _build_report_query_excel(
        start_date=start_date,
        end_date=end_date,
        shift_id=shift_id,
        platform_id=platform_id,
        db=db,
        current_user=current_user,
    )

    SHARED_REPORT_DIR.mkdir(parents=True, exist_ok=True)
    token = uuid4().hex
    stored_filename = f"{token}.xlsx"
    file_path = SHARED_REPORT_DIR / stored_filename
    file_path.write_bytes(output.getvalue())

    share = ReportExcelShare(
        token=token,
        tenant_id=get_current_tenant_id(db, current_user),
        created_by=current_user.id,
        filename=filename,
        file_path=str(file_path),
        start_date=date.fromisoformat(start_date),
        end_date=date.fromisoformat(end_date),
        expires_at=beijing_now() + timedelta(days=7),
    )
    db.add(share)
    db.add(
        AuditLog(
            user_id=current_user.id,
            module="report_share",
            action="create_excel_share",
            before_data=None,
            after_data=f"token={token},start={start_date},end={end_date}",
        )
    )
    db.commit()
    return {
        "token": token,
        "filename": filename,
        "url": f"/api/public/report-excel/{token}",
        "expires_at": share.expires_at,
    }


@PUBLIC_ROUTER.get("/report-excel/{token}")
def download_shared_report_excel(token: str, db: Session = Depends(get_db)):
    share = db.execute(select(ReportExcelShare).where(ReportExcelShare.token == token)).scalar_one_or_none()
    if share is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="分享链接不存在")
    if share.expires_at < beijing_now():
        raise HTTPException(status_code=status.HTTP_410_GONE, detail="分享链接已过期")

    file_path = Path(share.file_path)
    if not file_path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="分享文件不存在")
    return FileResponse(
        path=file_path,
        media_type=EXCEL_MEDIA_TYPE,
        filename=share.filename,
    )


@router.get("/excel")
def export_excel(
    year: int,
    month: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_module("reports.monthly")),
):
    allowed = _scoped_platform_ids(db, current_user)
    if allowed == []:
        allowed = [-1]
    wb = Workbook()

    ws_tx = wb.active
    ws_tx.title = "流水明细"
    ws_tx.append(["日期", "班次", "平台", "类型", "项目", "金额", "备注", "业务组"])
    tx_rows = db.execute(
        select(Transaction).where(
            extract("year", Transaction.bill_date) == year,
            extract("month", Transaction.bill_date) == month,
            Transaction.deleted_at.is_(None),
            Transaction.platform_id.in_(allowed) if allowed is not None else True,
        )
    ).scalars().all()
    for tx in tx_rows:
        ws_tx.append(
            [
                tx.bill_date.isoformat(),
                tx.shift_id,
                tx.platform_id,
                tx.type,
                tx.category_id,
                float(tx.amount),
                tx.remark,
                tx.biz_group_no,
            ]
        )
    apply_money_format(ws_tx, [6])

    ws_daily = wb.create_sheet("每日汇总")
    ws_daily.append(["日期", "班次", "平台", "收入", "支出", "净盈利"])
    daily_rows = db.execute(
        select(DailySummary).where(
            extract("year", DailySummary.bill_date) == year,
            extract("month", DailySummary.bill_date) == month,
            DailySummary.platform_id.in_(allowed) if allowed is not None else True,
        )
    ).scalars().all()
    for row in daily_rows:
        ws_daily.append(
            [
                row.bill_date.isoformat(),
                row.shift_id,
                row.platform_id,
                float(row.total_income),
                float(row.total_expense),
                float(row.net_profit),
            ]
        )
    apply_money_format(ws_daily, [4, 5, 6])

    ws_bal = wb.create_sheet("账户余额")
    ws_bal.append(["日期", "班次", "账户", "理论余额", "实际余额", "差额"])
    bal_rows = db.execute(
        select(AccountSnapshot).where(
            extract("year", AccountSnapshot.bill_date) == year,
            extract("month", AccountSnapshot.bill_date) == month,
        )
    ).scalars().all()
    for row in bal_rows:
        ws_bal.append(
            [
                row.bill_date.isoformat(),
                row.shift_id,
                row.account_id,
                float(row.theoretical_balance),
                float(row.actual_balance) if row.actual_balance is not None else None,
                float(row.difference),
            ]
        )
    apply_money_format(ws_bal, [4, 5, 6])

    ws_month = wb.create_sheet("月度汇总")
    ws_month.append(["月份", "总收入", "总支出", "净盈利"])
    monthly_row = db.execute(
        select(func.sum(DailySummary.total_income), func.sum(DailySummary.total_expense), func.sum(DailySummary.net_profit)).where(
            extract("year", DailySummary.bill_date) == year,
            extract("month", DailySummary.bill_date) == month,
            DailySummary.platform_id.in_(allowed) if allowed is not None else True,
        )
    ).first()
    ws_month.append([f"{year:04d}-{month:02d}", float(monthly_row[0] or 0), float(monthly_row[1] or 0), float(monthly_row[2] or 0)])
    apply_money_format(ws_month, [2, 3, 4])

    ws_platform = wb.create_sheet("平台汇总")
    ws_platform.append(["平台ID", "收入", "支出", "净盈利"])
    agg_rows = db.execute(
        select(Transaction.platform_id, Transaction.type, Transaction.amount).where(
            extract("year", Transaction.bill_date) == year,
            extract("month", Transaction.bill_date) == month,
            Transaction.deleted_at.is_(None),
            Transaction.platform_id.in_(allowed) if allowed is not None else True,
        )
    ).all()
    plat_data: dict[int, dict[str, float]] = {}
    for platform_id, tx_type, amount in agg_rows:
        if platform_id not in plat_data:
            plat_data[platform_id] = {"income": 0.0, "expense": 0.0}
        if tx_type == "income":
            plat_data[platform_id]["income"] += float(amount or 0)
        elif tx_type == "expense":
            plat_data[platform_id]["expense"] += float(amount or 0)
    for pid, val in plat_data.items():
        ws_platform.append([pid, val["income"], val["expense"], val["income"] - val["expense"]])
    apply_money_format(ws_platform, [2, 3, 4])

    ws_project = wb.create_sheet("项目汇总")
    ws_project.append(["项目ID", "类型", "金额"])
    project_rows = db.execute(
        select(Transaction.category_id, Transaction.type, func.sum(Transaction.amount)).where(
            extract("year", Transaction.bill_date) == year,
            extract("month", Transaction.bill_date) == month,
            Transaction.deleted_at.is_(None),
            Transaction.platform_id.in_(allowed) if allowed is not None else True,
        ).group_by(Transaction.category_id, Transaction.type)
    ).all()
    for r in project_rows:
        ws_project.append([r[0], r[1], float(r[2] or 0)])
    apply_money_format(ws_project, [3])

    ws_logs = wb.create_sheet("修改记录")
    ws_logs.append(["用户ID", "模块", "动作", "修改前", "修改后", "时间"])
    logs = db.execute(select(AuditLog).order_by(AuditLog.id.desc()).limit(2000)).scalars().all()
    for log in logs:
        ws_logs.append([log.user_id, log.module, log.action, log.before_data, log.after_data, log.created_at.isoformat()])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"accounting-{year:04d}-{month:02d}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/super-tenant-summary-excel")
def export_super_tenant_summary_excel(
    start_date: str | None = None,
    end_date: str | None = None,
    tenant_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles({UserRole.SUPER_ADMIN.value, UserRole.PLATFORM_VIEWER.value})),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "租户汇总"
    ws.append(["租户ID", "租户名称", "总收入", "总支出", "净利润"])

    stmt = (
        select(
            Tenant.id,
            Tenant.name,
            func.sum(DailySummary.total_income),
            func.sum(DailySummary.total_expense),
            func.sum(DailySummary.net_profit),
        )
        .select_from(Tenant)
        .join(TenantPlatformAccess, TenantPlatformAccess.tenant_id == Tenant.id)
        .join(DailySummary, DailySummary.platform_id == TenantPlatformAccess.platform_id)
    )
    if start_date:
        stmt = stmt.where(DailySummary.bill_date >= start_date)
    if end_date:
        stmt = stmt.where(DailySummary.bill_date <= end_date)
    if tenant_id:
        stmt = stmt.where(Tenant.id == tenant_id)
    rows = db.execute(stmt.group_by(Tenant.id, Tenant.name).order_by(Tenant.id.asc())).all()

    total_income = 0.0
    total_expense = 0.0
    total_net = 0.0
    for r in rows:
        income = float(r[2] or 0)
        expense = float(r[3] or 0)
        net = float(r[4] or 0)
        total_income += income
        total_expense += expense
        total_net += net
        ws.append([int(r[0]), r[1], income, expense, net])

    ws.append(["总计", "-", total_income, total_expense, total_net])
    apply_money_format(ws, [3, 4, 5])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = "super-tenant-summary.xlsx"
    db.add(
        AuditLog(
            user_id=current_user.id,
            module="exports",
            action="export_super_tenant_summary_excel",
            after_data=f"start_date={start_date},end_date={end_date},tenant_id={tenant_id}",
        )
    )
    db.commit()
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
